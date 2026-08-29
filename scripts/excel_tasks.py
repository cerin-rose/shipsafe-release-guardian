#!/usr/bin/env python3
"""Read FINAL_REPORT.md and write ShipSafe_Tasks.xlsx with one row per check."""
import re, pathlib, subprocess
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT          = pathlib.Path(__file__).parent.parent
REPORT        = ROOT / "FINAL_REPORT.md"
OUT           = ROOT / "ShipSafe_Tasks.xlsx"
DASHBOARD_URL = "http://127.0.0.1:5001"

SEVERITY_FILL = {
    "CRITICAL": "FFCCCC", "HIGH": "FFE0CC", "MEDIUM": "FFFACC", "LOW": "E8F4E8",
}
STATUS_ICON = {"✅": "PASS", "🔴": "FAIL", "🟠": "FAIL", "🟡": "WARN"}

# ── Actionable fix hints (What Fix column) ────────────────────────────────────
FIX_HINTS = {
    "C02": "Upgrade or replace the vulnerable dependency flagged in pom.xml",
    "C04": "Add VAULT_ID=<id> to application.properties or environment config",
    "C05": "Remove hardcoded secret; inject via environment variable or vault",
    "C06": "Set management.endpoints.web.exposure.include=health in application.properties",
    "C07": "Add server.ssl.* properties and configure a keystore for HTTPS",
    "C08": "Add spring-boot-starter-security to pom.xml and create a SecurityFilterChain bean",
    "C09": "Add flyway-core or liquibase-core to pom.xml for managed schema migrations",
    "C14": "Add CUDA_VISIBLE_DEVICES env var to pipeline.yaml GPU job definition",
    "C15": "Replace spring-boot-starter-webmvc with spring-boot-starter-web in pom.xml",
    "C16": "Declare VAULT_ID in the env block of deploy/codeengine.yaml",
}

# ── Created From Why: why each check exists ───────────────────────────────────
WHY_HINTS = {
    "C01": "API compatibility gate — prevents breaking changes reaching production consumers",
    "C02": "Supply-chain security — known CVE in a direct dependency blocks safe deployment",
    "C03": "Python dependency audit — CVE in requirements.txt or lockfile poses runtime risk",
    "C04": "Secrets management — VAULT_ID must be bound before the app can reach the vault",
    "C05": "Credential hygiene — hardcoded secrets in config files leak via version control",
    "C06": "Attack surface — wildcard actuator exposure allows unauthenticated admin access",
    "C07": "Transport security — plain HTTP in production exposes credentials and data in transit",
    "C08": "Authentication gate — no security dependency means all endpoints are unprotected",
    "C09": "Data integrity — unmanaged schema changes risk migration failures on deploy",
    "C10": "Container hygiene — ports below 1024 require root privileges in the container",
    "C11": "Infrastructure hygiene — hardcoded S3 bucket couples pipeline to a specific account",
    "C12": "Notebook security — api_key literals in notebooks leak credentials via git history",
    "C13": "Notebook portability — hardcoded local paths break notebook execution in CI/CD",
    "C14": "ML reproducibility — missing CUDA env var causes silent CPU fallback on GPU nodes",
    "C15": "Dependency hygiene — phantom artifact fails resolution and blocks the build",
    "C16": "Deploy-time secrets — VAULT_ID absent in codeengine.yaml means app cannot start",
}


# ── git helpers ───────────────────────────────────────────────────────────────

def _git_repo_for(abs_path: pathlib.Path):
    """Return the nearest git repo root that tracks abs_path, plus the relative path within it."""
    candidate = abs_path.parent
    while candidate != candidate.parent:
        if (candidate / ".git").exists():
            return candidate, abs_path.relative_to(candidate)
        candidate = candidate.parent
    return None, None


def git_blame_author(file_path: str, line: str = "") -> str:
    """Return the author name from git blame for the given file (and optional line)."""
    abs_path = (ROOT / file_path).resolve()
    if not abs_path.exists():
        return "N/A"
    repo_root, rel = _git_repo_for(abs_path)
    if repo_root is None:
        return "N/A"
    try:
        cmd = ["git", "blame", "--porcelain"]
        if line.isdigit():
            cmd += [f"-L{line},{line}"]
        cmd.append(str(rel))
        out = subprocess.check_output(cmd, cwd=str(repo_root), stderr=subprocess.DEVNULL, text=True)
        for l in out.splitlines():
            if l.startswith("author "):
                author = l[7:].strip()
                return author if author != "Not Committed Yet" else "Uncommitted"
        return "Unknown"
    except subprocess.CalledProcessError:
        return "N/A"


def git_file_date(file_path: str) -> str:
    """Return the last-commit date for the file (ISO format)."""
    abs_path = (ROOT / file_path).resolve()
    if not abs_path.exists():
        return ""
    repo_root, rel = _git_repo_for(abs_path)
    if repo_root is None:
        return ""
    try:
        out = subprocess.check_output(
            ["git", "log", "-1", "--format=%ci", "--", str(rel)],
            cwd=str(repo_root), stderr=subprocess.DEVNULL, text=True,
        ).strip()
        # "2024-03-15 10:22:01 +0000" → "2024-03-15 10:22"
        return out[:16] if out else ""
    except subprocess.CalledProcessError:
        return ""


# ── Table parser ──────────────────────────────────────────────────────────────

def parse_table(text: str):
    rows = []
    # Parse domain from the table line (column 3)
    for line in text.splitlines():
        m = re.match(
            r"\|\s*(C\d+)\s*\|\s*(\w+)\s*\|\s*(\w+)\s*\|\s*(.+?)\s*\|\s*`(.+?)`\s*\|\s*(\S+)\s*\|",
            line,
        )
        if not m:
            continue
        tid, sev, domain, bug, file_line, icon = m.groups()
        fp  = file_line.split(":")[0].strip()
        ln  = file_line.split(":")[1].strip() if ":" in file_line else ""
        status = STATUS_ICON.get(icon, "PASS")
        fix = (
            FIX_HINTS.get(tid, "Review finding and remediate")
            if status != "PASS"
            else "No action required — check passed"
        )
        rows.append({
            "id":        tid,
            "sev":       sev,
            "domain":    domain,
            "bug":       bug.strip(),
            "file_path": fp,
            "line":      ln,
            "file_line": file_line,
            "status":    status,
            "fix":       fix,
        })
    return rows


def parse_details(text: str) -> dict:
    """Extract the 'Finding Details' narrative per check ID."""
    details = {}
    current_id = None
    lines_buf  = []
    for line in text.splitlines():
        h = re.match(r"^###\s*(C\d+)\s*—\s*.+", line)
        if h:
            if current_id and lines_buf:
                details[current_id] = " ".join(lines_buf).strip()
            current_id = h.group(1)
            lines_buf  = []
        elif current_id and line.startswith("**Severity"):
            pass   # skip the bold meta line
        elif current_id and line.strip() and not line.startswith("#"):
            lines_buf.append(line.strip())
    if current_id and lines_buf:
        details[current_id] = " ".join(lines_buf).strip()
    return details


# ── Helpers ───────────────────────────────────────────────────────────────────

def due_date(sev: str) -> str:
    today = datetime.now().date()
    if sev == "CRITICAL":
        return today.strftime("%Y-%m-%d")
    if sev == "HIGH":
        return (today + timedelta(days=1)).strftime("%Y-%m-%d")
    return ""


def fix_link_formula(file_line: str) -> str:
    fp       = file_line.split(":")[0].strip()
    abs_path = (ROOT / fp).resolve()
    return f'=HYPERLINK("file:///{abs_path}", "Open")'


# ── Parse report ─────────────────────────────────────────────────────────────
text = REPORT.read_text(errors="ignore")
rows = parse_table(text)
details_map = parse_details(text)

generated    = ""
overall_scan = "UNKNOWN"
g = re.search(r"\*\*Generated:\*\*\s*([\d\-T: ]+)", text)
if g: generated = g.group(1).strip()
m = re.search(r"\*\*Overall:\*\*\s*(.+)", text)
if m: overall_scan = m.group(1).strip()

now_str       = datetime.now().strftime("%Y-%m-%d %H:%M")
today_str     = datetime.now().strftime("%Y-%m-%d")
pending_count = sum(1 for r in rows if r["status"] != "PASS")

# ── Workbook ──────────────────────────────────────────────────────────────────
if OUT.exists():
    wb = openpyxl.load_workbook(OUT)
    if "ShipSafe Tasks" in wb.sheetnames:
        del wb["ShipSafe Tasks"]
    ws = wb.create_sheet("ShipSafe Tasks", 0)
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ShipSafe Tasks"

# ── Style constants ───────────────────────────────────────────────────────────
BROWN       = "5C4B37"
ACCENT      = "8B7355"
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D6CBB8"),
    right= Side(style="thin", color="D6CBB8"),
)

def meta_fill():
    return PatternFill("solid", fgColor=BROWN)

def mk_font(bold=False, color="161616", size=10, underline=None):
    return Font(bold=bold, color=color, size=size, name="Calibri",
                underline=underline)

# ── Row 1: Meta banner (spans all 17 columns) ─────────────────────────────────
META = [
    (1,  "ShipSafe Release Guardian",            mk_font(bold=True, color="FFFFFF", size=13)),
    (2,  f"Report: {generated or today_str}",    mk_font(color="F5F1E8")),
    (3,  f"Run: {now_str}",                      mk_font(color="F5F1E8")),
    (4,  f"Pending fixes: {pending_count}",      mk_font(bold=True, color="FFCCCC")),
    (5,  f"Verdict: {overall_scan}",             mk_font(bold=True, color="FFCCCC" if "NO-GO" in overall_scan else "CCFFCC")),
    (17, f'=HYPERLINK("{DASHBOARD_URL}","Dashboard")', mk_font(bold=True, color="FFFFFF", underline="single")),
]
for col, val, fnt in META:
    c = ws.cell(1, col, val)
    c.font      = fnt
    c.fill      = meta_fill()
    c.alignment = Alignment(horizontal="left" if col < 17 else "center", vertical="center")

for col in range(1, 18):                         # fill all 17 banner cells
    c = ws.cell(1, col)
    if c.fill.fgColor.rgb == "00000000":
        c.fill = meta_fill()
ws.row_dimensions[1].height = 22

# ── Row 2: Column headers ─────────────────────────────────────────────────────
# Columns A–J: existing | K–Q: new
COL_HEADERS = [
    # A          B       C       D
    "Task ID",  "File", "Line", "Bug",
    # E            F                  G        H
    "Severity", "Fix Suggestion",   "Status", "Date",
    # I           J
    "Due Date", "Fix Link",
    # K              L                      M                N
    "Bug Number", "Person In Charge",    "Date Time",    "Details",
    # O             P           Q
    "What Fix",  "Pending",  "Created From Why",
]
ws.append(COL_HEADERS)
HDR_ROW = 2
for cell in ws[HDR_ROW]:
    cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=ACCENT)
    cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    cell.border    = THIN_BORDER
ws.row_dimensions[HDR_ROW].height = 18

# Column widths A–Q
WIDTHS = [
    9,  40,  6,  44,                  # A–D
    11, 52,   9,  18,                  # E–H
    12, 10,                            # I–J
    12, 22,  18,  40,                  # K–N
    46, 10,  52,                       # O–Q
]
for i, w in enumerate(WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Data rows ─────────────────────────────────────────────────────────────────
for r in rows:
    tid       = r["id"]
    fp        = r["file_path"]
    ln        = r["line"]
    bug       = r["bug"]
    sev       = r["sev"]
    domain    = r["domain"]
    fix       = r["fix"]
    status    = r["status"]
    file_line = r["file_line"]

    pic        = git_blame_author(fp, ln)
    file_date  = git_file_date(fp)
    detail     = details_map.get(tid, "No additional detail recorded")
    what_fix   = FIX_HINTS.get(tid, fix)
    pending    = "Yes" if status != "PASS" else "No"
    why        = WHY_HINTS.get(tid, f"{domain} — {bug}")
    bug_number = f"BUG-{tid}"
    task_status = "Pending" if status != "PASS" else "OK"

    ws.append([
        tid, fp, ln, bug, sev, fix, task_status, now_str,   # A–H
        due_date(sev),                                        # I
        fix_link_formula(file_line),                         # J
        bug_number,                                           # K
        pic,                                                  # L
        file_date or now_str,                                 # M
        detail,                                               # N
        what_fix,                                             # O
        pending,                                              # P
        why,                                                  # Q
    ])
    row_idx = ws.max_row
    fill = PatternFill("solid", fgColor=SEVERITY_FILL.get(sev, "FFFFFF"))
    for col in range(1, 18):
        cell = ws.cell(row_idx, col)
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        # J (Fix Link) and P (Pending) stay white; all others get severity fill
        if col not in (10,):
            cell.fill = fill
        if col == 16:    # P: Pending — bold + colour
            cell.font = Font(
                bold=True,
                color="A21918" if pending == "Yes" else "1E6B2E",
                size=10, name="Calibri",
            )
        if col == 10:    # J: Fix Link — centre
            cell.alignment = Alignment(horizontal="center", vertical="top")

ws.freeze_panes = "A3"

# ── History sheet ─────────────────────────────────────────────────────────────
HIST = "History"
if HIST not in wb.sheetnames:
    wh = wb.create_sheet(HIST)
    wh.append(["Date", "Overall", "Pass", "Fail", "Warn", "Total"])
    for cell in wh[1]:
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=BROWN)
        cell.alignment = Alignment(horizontal="center")
    for col, w in zip("ABCDEF", [18, 36, 8, 8, 8, 8]):
        wh.column_dimensions[col].width = w
else:
    wh = wb[HIST]

pass_c = sum(1 for r in rows if r["status"] == "PASS")
fail_c = sum(1 for r in rows if r["status"] == "FAIL")
warn_c = sum(1 for r in rows if r["status"] == "WARN")
wh.append([now_str, overall_scan, pass_c, fail_c, warn_c, len(rows)])

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Written {len(rows)} tasks → {OUT.name}")
print(f"  Columns A–Q  : {len(COL_HEADERS)}")
print(f"  Pending fixes: {pending_count}")
print(f"  Pass / Fail / Warn: {pass_c} / {fail_c} / {warn_c}")
print(f"  History rows : {wh.max_row - 1}")
